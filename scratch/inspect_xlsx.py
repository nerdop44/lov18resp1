import openpyxl

def inspect_file(filepath):
    print(f"--- Inspecting {filepath} ---")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print("Sheets:", wb.sheetnames)
    sheet = wb.active
    print("Max Row:", sheet.max_row, "Max Column:", sheet.max_column)
    
    # Print the first 25 rows, first 22 columns if they have content
    for r in range(1, min(40, sheet.max_row + 1)):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(25, sheet.max_column + 1))]
        if any(x is not None for x in row_vals):
            # Format numbers nicely
            formatted = []
            for val in row_vals:
                if isinstance(val, float):
                    formatted.append(f"{val:.2f}")
                else:
                    formatted.append(str(val))
            # Find the index of row_vals to show
            print(f"Row {r:02d}: " + " | ".join(formatted[:22]))

print("Libro de compra (10):")
try:
    inspect_file("/home/nerdop/Descargas/Libro_de_compra (10).xlsx")
except Exception as e:
    import traceback
    traceback.print_exc()

